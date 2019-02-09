import csv
import datetime
import os.path
import sys
import xml.etree.ElementTree
from typing import List, Dict, Optional, Any

import xlrd
from openpyxl import Workbook

from .Arrays import Arrays
from .Files import Files

ExRows = List[List[str]]


class Excel:
    default_sheet_name: str = 'Sheet'
    workbook: Workbook
    file_name: str
    row_counts: Dict[str, int]

    def __init__(self, file_name: str, default_sheet_name: str = None, save_on_init=True):
        self.workbook = Workbook()
        self.row_counts = {}

        if default_sheet_name is not None:
            self.workbook.active.title = default_sheet_name
            self.row_counts[default_sheet_name] = 0
        else:
            self.row_counts[self.default_sheet_name] = 0

        self.file_name = file_name

        if save_on_init:
            self.workbook.save(filename=self.file_name)

    @staticmethod
    def separate_excel_and_cvs_files():
        """ postc: returns a list of all the excel (xlsx) and csv files in the
                    current working directory
            list[0] = excel files, list[1] = csv files
            """
        try:
            main_module_path = os.path.dirname(os.path.realpath(__file__))
        except NameError:
            main_module_path = os.path.dirname(os.path.abspath(sys.argv[0]))
        files = Files.get_all_files_(main_module_path)
        return [[x for x in files if x[-4:] == 'xlsx'], [x for x in files if x[-3:] == 'csv']]

    @staticmethod
    def merge_sheets_remove_duplicates(file_name, duplicates, uid):
        """ prec: file_name is a file name, uid is a string of the column that is the unique
                    identifier
            postc: creates a spreadsheet file with rows of all the spreadsheets in the current
                    working directory. Basically a merger function for excel and csv files"""
        print("Performing operations on all XLSX and CSV files in current folder")

        # get files and create list master list for all rows
        files = Excel.separate_excel_and_cvs_files()
        master_list = []
        if duplicates:
            master_list = [["File"], {}]
        else:
            master_list.append([])

        # loop through excel files
        for name in files[0]:
            print("Merging: " + name)
            Excel.merge_excel_files(name, master_list, duplicates, uid)

        # then loop through csv files
        for name in files[1]:
            print("Merging: " + name)
            Excel.merge_cvs_files(name, master_list, duplicates, uid)

        if duplicates:
            Excel.bring_uid_to_front(master_list[0], uid)
            master_list = Excel.convert_excel_dictionaries(master_list)

        Excel.create_master_sheet(file_name + ".xls", master_list)

    def save_rows(self, rows: ExRows, sheet_name: str = None, row_start: int = 0):
        sheet = self.get_sheet(sheet_name)
        sheet_name = sheet.title

        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row):
                datum = Excel.remove_illegal_chars(value)
                sheet.cell(column=col_idx + 1, row=row_idx + 1 + row_start, value=datum)

        self.row_counts[sheet_name] = row_start + len(rows)
        self.workbook.save(filename=self.file_name)

    def get_row_count(self, sheet_name: str = None):
        sheet = self.get_sheet(sheet_name)
        return self.row_counts[sheet.title]

    def add_rows(self, rows: ExRows, sheet_name: str = None):
        sheet = self.get_sheet(sheet_name)
        sheet_name = sheet.title

        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row):
                datum = Excel.remove_illegal_chars(value)
                sheet.cell(column=col_idx + 1, row=row_idx + 1 + self.row_counts[sheet_name],
                           value=datum)

        self.row_counts[sheet_name] += len(rows)
        self.workbook.save(filename=self.file_name)

    def add_sheet(self, sheet_name: str):
        if sheet_name not in self.row_counts:
            self.workbook.create_sheet(title=sheet_name)
            self.row_counts[sheet_name] = 0

    def get_sheet(self, sheet_name: str):
        if sheet_name is None and len(self.row_counts) == 1:
            return self.workbook.active
        if sheet_name is not None:
            return self.workbook[sheet_name]

        raise KeyError('There was no sheet name found from the given one. '
                       'The sheet_name must be specified if there are multiple '
                       'sheets in the spreadsheet')

    @staticmethod
    def create_master_sheet(file_name, master_list):
        """ prec: file_name is the name of the output file for the merged spreadsheet,
                    master_list is the list of all the rows ready to be saved to the spreadsheet
            postc: merges the new rows from the excel file into the row dictionary.
                    Removes duplicates
        """
        print("Creating spreadsheet file")
        master = Workbook()
        sheet = master.active

        for row_idx, row in enumerate(master_list):
            for col_idx, value in enumerate(row):
                datum = Excel.remove_illegal_chars(value)
                sheet.cell(column=col_idx + 1, row=row_idx + 1, value=datum)

        print("Saving")
        master.save(filename=file_name)

    @staticmethod
    def convert_json_list_to_rows(json_list: List[dict],
                                  headers: Optional[List[str]] = None) -> ExRows:
        if not json_list:
            return []
        output: ExRows = []

        headers = list(json_list[0].keys()) if headers is None else headers
        output.append(headers)
        for jsons in json_list:
            new_row: List[str] = []
            for header in headers:
                datum: str = str(jsons.get(header))
                new_row.append(Excel.remove_illegal_chars(datum))
            output.append(new_row)

        return output

    @staticmethod
    def write_errors_to_file_helper(file_name, errors):
        """ prec: errors is a specialized error array. Holds all the cells that
                    couldn't be saved to the spreadsheet
                    Each data item has 3 items - row #, column #, and
                    the cell date (in the order per item) file_name is the file to save it to
            postc: creates a text file to hold the errors and outputs to
                    a text file ('Errors.txt')
        """
        if errors:
            print("Creating error file")
            error_file = open(file_name + ".txt", "w")
            error_file.write(
                "Was not able to save the following cells to their respective spreadsheets:\n")
            for error in errors:
                error_file.write("Row: " + error[1] + " Column: " +
                                 error[2] + " in file " + error[0] + "\n")
                error_file.write(errors[3])

    @staticmethod
    def bring_uid_to_front(header_list, uid):
        """ prec: headers is a list of strings and uid is a string
            postc: performs the swap to put the uid to the front of the column list"""
        Arrays.swap(header_list, 0, Excel.get_uid_index(header_list, uid))

    @staticmethod
    def merge_new_header_columns(headers):
        for idx, header in enumerate(headers):
            header = header.lower()
            if header == "university":
                headers[idx] = "School"
            if "course" in header:
                headers[idx] = "Course"
            if header == "prof":
                headers[idx] = "Professor"
            if "first" in header:
                headers[idx] = "First Name"
            if "last" in header:
                headers[idx] = "Last Name"

    # helper function for merge_excel_files and merge_csv_files
    @staticmethod
    def update_headers_get_uid_index(keys_list, new_header_row, uid):
        """ prec: keys_list and new_header_row are lists, uid is a string
            postc: 1) merges duplicate column names in the new header list new_header_row
                    2) merges the column names in new_header_row with the columns in
                        keys_list - removes duplicates
                    3) returns the index of the uid
        """
        Excel.merge_new_header_columns(new_header_row)
        Excel.merge_new_header(keys_list, new_header_row)

        return Excel.get_uid_index(new_header_row, uid)

    @staticmethod
    def merge_excel_files(file_name, master_list, duplicates, uid):
        """ prec: file_name is the name of the file being merged, keys_list is a list of the
                    column names, row_dictionary is the dictionary file for all the rows,
                    and uid is the unique identifier as a string
            postc: merges the new rows from the excel file into the row dictionary.
                 Removes duplicates
        """
        workbook = xlrd.open_workbook(file_name)  # open excel file
        sheets = workbook.sheets()  # get sheet

        # loop through sheets
        for sheet in sheets:
            # get headers and set email index
            headers = sheet.row_values(0)
            if duplicates:
                uid_idx = Excel.update_headers_get_uid_index(master_list[0], headers, uid)
                if uid_idx < 0:
                    return

                # loop through rows in sheet
                for row_num in range(1, sheet.nrows):
                    Excel.merge_file_without_duplicates(file_name, master_list[1], headers,
                                                        sheet.row_values(row_num), uid_idx)
            else:
                if not master_list[0]:
                    master_list[0] += headers
                    master_list[0].append("File")

                # loop through rows in sheet
                for row_num in range(1, sheet.nrows):
                    Excel.merge_file_with_duplicates(file_name, master_list,
                                                     sheet.row_values(row_num))

    @staticmethod
    def merge_cvs_files(file_name, master_list, duplicates, uid):
        """ prec: file_name is the name of the file being merged, keys_list is a
                    list of the column names, row_dictionary is the dictionary file
                    for all the rows, and uid is the unique identifier as a string
            postc: merges the new rows from the csv file into the row dictionary.
                    Removes duplicates
        """
        with open(file_name, 'rU', encoding='ISO-8859-1') as csv_file:
            reader = csv.reader(csv_file, dialect='excel')
            index = 0
            uid_idx = 0
            headers = []

            for current_row in reader:
                if duplicates:
                    if index == 0:
                        headers = current_row
                        uid_idx = Excel.update_headers_get_uid_index(master_list[0], headers, uid)
                        if uid_idx < 0:
                            return

                        index += 1
                    else:
                        Excel.merge_file_without_duplicates(file_name, master_list[1],
                                                            headers, current_row,
                                                            uid_idx)
                else:
                    if index == 0:
                        headers = current_row

                        if not master_list[0]:
                            master_list[0] += headers
                            master_list[0].append("File")

                        index += 1
                    else:
                        Excel.merge_file_with_duplicates(file_name, master_list, current_row)

    @staticmethod
    def merge_file_with_duplicates(file_name, rows_list, current_row):
        """ prec: file_name is the name of the file being merged, row_list is the list of rows,
                        and current_row is the current row to be merged with the list
            postc: helper merger function for merging both excel and csv files. Keeps duplicates"""
        new_row = []

        # loop through cells in row
        for new_cell in enumerate(current_row):
            if not isinstance(new_cell, str):
                new_cell = str(new_cell)
            new_row.append(new_cell)

        new_row.append(file_name)
        rows_list.append(new_row)

    @staticmethod
    def merge_file_without_duplicates(file_name, row_dictionary, headers, current_row, uid):
        """ prec: file_name is the name of the file being merged, headers is a list of
                    the column names, current_row is the current row being merged,
                    row_dictionary is the dictionary file for all the rows,
                    and uid is the unique identifier as a string
                postc: helper merger function for merging both excel and csv files.
                Removes duplicates
        """
        # check if current row's email is in merged rows dictionary
        # if not, put it in
        uid_as_string = current_row[uid]
        if uid_as_string not in row_dictionary:
            temp = {"File": [file_name]}
            row_dictionary[uid_as_string] = temp

        # get said merged row
        merged_row = row_dictionary[uid_as_string]

        # current_row = current row in sheets
        # merged_row = new row in merged dictionary

        # loop through cells in row
        for cell_idx, cell in enumerate(current_row):
            # check if cell is not email index
            # ignoring because already added above ^^
            if cell_idx != uid:
                # check if current header in merged_row dictionary
                # if not, add it
                if headers[cell_idx] not in merged_row:
                    merged_row[headers[cell_idx]] = []

                if current_row[cell_idx] not in merged_row[headers[cell_idx]]:
                    merged_row[headers[cell_idx]].append(cell)

    @staticmethod
    def remove_illegal_chars(value: str):
        """ prec: value is a string
            postc: returns the string without unicode characters above index 128"""
        temp = []
        for item in enumerate(value):
            char: str = item[1]
            if ord(char) < 128:
                temp.append(char)

        output = ''.join(temp)

        if output == 'True':
            return 'Yes'
        if output == 'False':
            return 'No'
        if output == 'None':
            return ''

        return output

    @staticmethod
    def get_uid_index(headers, uid):
        """ prec: headers is a list of strings and uid is a string
            postc: returns the index of the unique identifier (uid), or -1 if it is not found"""
        return Arrays.get_index_in_list_insensitive(headers, uid)

    @staticmethod
    def merge_new_header(keys_list, headers):
        """ prec: keys_list and headers are lists
            postc: merges the two lists and removes duplicates"""
        Arrays.merge_lists_remove_duplicates(keys_list, headers)

    @staticmethod
    def convert_excel_dictionaries(master_list):
        """ prec: headers is a list of column names, rows is the row dictionary
                    with all the merged rows
            postc: returns a 2D array with all the rows for the spreadsheet"""
        headers = master_list[0]
        rows = master_list[1]

        new_master_list = [list(headers)]
        row_index = 0

        for row in rows:
            new_list = [''] * (len(headers) - 1)
            new_list[0] = row
            for column in rows[row]:
                cell_value_list = rows[row][column]

                cell_output = ""
                for cell in cell_value_list:
                    if not isinstance(cell, str):
                        cell = str(cell)
                    cell_output += cell + " | "

                cell_output = cell_output[:-3]
                new_list[headers.index(column)] = cell_output
            new_master_list.append(new_list)
            row_index += 1

        return new_master_list

    @staticmethod
    def excel_date_as_datetime(excel_date, date_mode):
        """ prec: excel_date is an excel date, date_mode is the date_mode
                    (0 for 1900-based, 1 for 1904 based)
            postc: returns date as a string
        """
        # date_mode: 0 for 1900-based, 1 for 1904-based
        return datetime.datetime(1899, 12, 30) + datetime.timedelta(
            days=excel_date + 1462 * date_mode)

    @staticmethod
    def open_file(file_path, args):
        """ prec: file is a valid path, args are valid file arguments
                    (e.g. 'r' for read, 'w' for write, etc)
            postc: returns the opened file"""
        in_pipe = open(file_path, args)
        stuff = in_pipe.read()
        return stuff

    # api: http://xlrd.readthedocs.io/en/latest/api.html
    @staticmethod
    def copy_excel_file_to_cvs(output_path: str, filename: str):
        """ prec: file is a valid excel file path
            postc: creates a csv file from the given excel file"""
        workbook = xlrd.open_workbook(output_path + '/' + filename)
        sheets = workbook.sheets()
        i = 1

        for sheet in sheets:
            your_csv_file = open('{}/{}_{}.csv'.format(output_path, filename, str(i)), 'w')
            csv_writer = csv.writer(your_csv_file, delimiter=",", quoting=csv.QUOTE_ALL)

            for row_num in range(sheet.nrows):
                csv_writer.writerow(sheet.row_values(row_num))

            your_csv_file.close()
            i += 1

    @staticmethod
    def convert_xml_to_excel(xml_file, new_file_name):
        """ prec: xml_file is an xml file, new_file_name is the name of the
                    spreadsheet to be created
            postc: creates an excel file from the given xml file (tags are columns, text is rows)
        """
        master_list = [[]]
        root = xml.etree.ElementTree.parse(xml_file).getroot()
        for child in root[0]:
            master_list[0].append(child.tag)
        for child in root:
            new_row = []
            for tag in child:
                new_row.append(str(tag.text))
            master_list.append(new_row)

        Excel.create_master_sheet(master_list, new_file_name)

    @staticmethod
    def convert_csv_file_to_rows(csv_file_name: str) -> ExRows:
        output: ExRows = []
        with open(csv_file_name, 'r') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                output.append(row)

        return output

    @staticmethod
    def convert_csv_to_json_via_text(csv_file_name, encoding: Optional[str] = 'utf-8',
                                     delimiter: Optional[str] = ',') -> List[Dict[str, Any]]:
        json_list: List[Dict[str, Any]] = []

        with open(csv_file_name, 'r', encoding=encoding) as csv_file:
            print(csv_file.read())
            lines: List[str] = csv_file.read().split('\n')
            headers: List[str] = lines.pop(0).split(delimiter)

            for line in lines:
                if line.strip() == '':
                    continue

                new_json_list: Dict[str, Any] = {}

                cells: List[str] = line.split(delimiter)
                for idx, cell in enumerate(cells):
                    new_json_list[headers[idx]] = cell.replace('"', '').replace('\'', '')

                json_list.append(new_json_list)
        return json_list

    @staticmethod
    def convert_csv_file_to_json_list(csv_file_name, encoding: Optional[str] = 'utf-8',
                                      delimiter: Optional[str] = ',') -> List[Dict[str, Any]]:
        """
        Takes a csv file and converts it to a json list
        :param delimiter:
        :param encoding:
        :param csv_file_name: The CSV file to convert
        :return: A json representation of the CSV file
        """
        json_list: List[Dict[str, Any]] = []

        with open(csv_file_name, 'r', encoding=encoding) as csv_file:
            reader = csv.reader(csv_file, delimiter=delimiter)
            headers = []
            row_idx = 0
            for row in reader:
                if row_idx == 0:
                    headers = row
                    row_idx += 1
                else:
                    new_json_list: Dict[str, Any] = {}

                    for cell_idx, cell in enumerate(row):
                        new_json_list[headers[cell_idx]] = cell

                    json_list.append(new_json_list)
        return json_list

    @staticmethod
    def convert_json_to_excel_rows(headers, json_list):
        """
        Takes a json dictionary and a list of headers and converts it
        to a list (for Excel formatting)

        :param headers: The headers of the Excel file
        :param json_list: A json dictionary
        :return: A list of Excel rows
        """
        master_list = [headers]

        for json_item in json_list:
            new_row = []
            for key in headers:
                if key in json_item:
                    new_row.append(json_item[key])
                else:
                    new_row.append('')
            master_list.append(new_row)

        return master_list
